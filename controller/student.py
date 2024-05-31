from io import BytesIO
from typing import List, Optional, Union, Literal
from fastapi.responses import StreamingResponse
from fastapi import APIRouter, HTTPException, UploadFile, File
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from prisma.errors import ForeignKeyViolationError

from controller.group import Group
from db.db import prisma
from pydantic import BaseModel
from datetime import datetime


class StudentCreate(BaseModel):
    name: str
    groupId: int
    status: bool


class StudentUpdate(StudentCreate):
    pass


class Student(StudentCreate):
    id: int
    createdAt: datetime
    updatedAt: datetime


class StudentWithGroup(Student):
    group: Group


class ChangeStatus(BaseModel):
    studentId: int
    status: bool


class StudentFilterOrder(BaseModel):
    by: Union[Literal['id'], Literal['name'], Literal['status'], Literal['groupId'], Literal['createdAt']]
    direction: Union[Literal['asc'], Literal['desc']]


class StudentFilter(BaseModel):
    skip: int = 0
    take: int = 20
    name: Optional[str] = None
    groupId: Optional[int] = None
    courseNumber: Optional[int] = None
    status: Optional[bool] = None
    order: Optional[StudentFilterOrder] = StudentFilterOrder(by='id', direction='asc')


class StudentList(BaseModel):
    students: List[StudentWithGroup]
    count: int
    skip: int
    take: int


router = APIRouter()


@router.post("/filter", response_model=StudentList)
async def get_students(filter: StudentFilter):
    where = {}
    groupWhere = {}
    order = {"id": "asc"}
    for key, value in filter.dict().items():
        if key in ["skip", "take"]:
            continue
        if value is None:
            continue

        if key == "courseNumber":
            groupWhere[key] = value

        if key == "name":
            where[key] = {"contains": value}
            continue

        if key == "order":
            order = {}
            order.update({value.get("by", "id"): value.get("direction", "asc")})
            continue

        where[key] = value
    count = await prisma.student.count(where=where)
    students = await prisma.student.find_many(
        take=filter.take,
        skip=filter.skip,
        where=where,
        order=order,
        include={
            'group': {
                'where': groupWhere
            }
        }
    )
    result = []
    for s in students:
        group = Group(id=s.group.id, groupNumber=s.group.groupNumber, courseNumber=s.group.courseNumber)
        result.append(StudentWithGroup(
            id=s.id,
            createdAt=s.createdAt,
            updatedAt=s.updatedAt,
            name=s.name,
            groupId=s.groupId,
            status=s.status,
            group=group
        ))
    return StudentList(students=result, count=count, skip=filter.skip, take=filter.take)


@router.post("/export")
async def export(filter: StudentFilter):
    students_list = await get_students(filter)
    if not students_list.students:
        raise HTTPException(status_code=400, detail="Empty student list")
    workbook = Workbook()
    groups = {}

    for student in students_list.students:
        group_name = student.group.groupNumber
        if group_name not in groups:
            groups[group_name] = []
        groups[group_name].append(student)

    for group_name, students in groups.items():
        worksheet = workbook.create_sheet(title=group_name)
        worksheet.append(["ID", "Name", "Status", "Created At", "Updated At"])
        for student in students:
            worksheet.append([
                student.id,
                student.name,
                student.status,
                student.createdAt.strftime("%Y-%m-%d %H:%M:%S"),
                student.updatedAt.strftime("%Y-%m-%d %H:%M:%S")
            ])

    workbook.remove(workbook["Sheet"])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={"Content-Disposition": "attachment; filename=students_export.xlsx"})


@router.post("/import")
async def import_students(file: UploadFile = File(...)):
    try:
        workbook = load_workbook(filename=BytesIO(await file.read()))
        async with prisma.batch_() as transaction:
            for sheet_name in workbook.sheetnames:
                group = await prisma.group.find_unique(where={
                    "groupNumber": sheet_name.upper()
                })
                if not group:
                    continue
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    student_id, name, status, created_at, updated_at = row
                    if not student_id:
                        student_id = -1

                    student = Student(
                        id=student_id,
                        name=name,
                        groupId=group.id,
                        status=status,
                        createdAt=datetime.now(),
                        updatedAt=datetime.now()
                    )
                    transaction.student.upsert(
                        where={"id": student.id},
                        data={
                            "update": {
                                "name": student.name,
                                "groupId": student.groupId,
                                "status": student.status,
                            },
                            "create": {
                                "name": student.name,
                                "groupId": student.groupId,
                                "status": student.status,
                            }
                        }
                    )
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/", response_model=Student)
async def create_student(student: StudentCreate):
    try:
        created_student = await prisma.student.create(
            data={
                "name": student.name,
                "groupId": student.groupId,
                "status": student.status
            }
        )
        return created_student
    except ForeignKeyViolationError as e:
        raise HTTPException(status_code=404, detail="Group not found")


@router.get("/{student_id}", response_model=Student)
async def read_student(student_id: int):
    student = await prisma.student.find_unique(where={"id": student_id})
    if student is None:
        raise HTTPException(status_code=404, detail="Student not found")
    return student


@router.patch("/changeStatus")
async def change_status(data: ChangeStatus):
    updated_student = await prisma.student.update(
        where={"id": data.studentId},
        data={
            "status": data.status
        }
    )
    if updated_student is None:
        raise HTTPException(status_code=404, detail="Student not found")
    return


@router.put("/{student_id}", response_model=Student)
async def update_student(student_id: int, student: StudentUpdate):
    updated_student = await prisma.student.update(
        where={"id": student_id},
        data={
            "name": student.name,
            "groupId": student.groupId,
            "status": student.status
        }
    )
    if updated_student is None:
        raise HTTPException(status_code=404, detail="Student not found")
    return updated_student


@router.delete("/{student_id}", response_model=Student)
async def delete_student(student_id: int):
    deleted_student = await prisma.student.delete(where={"id": student_id})
    if deleted_student is None:
        raise HTTPException(status_code=404, detail="Student not found")
    return deleted_student
